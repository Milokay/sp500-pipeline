"""
Excel report builder.
Generates a formatted 5-sheet Excel workbook from analysis results.
"""

import logging
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


def build_excel_report(results: list[dict], output_path: str) -> str:
    """
    Generate a formatted Excel workbook with 5 sheets.

    Args:
        results: List of dicts, each containing merged analysis data for one ticker
        output_path: Path to save the Excel file

    Returns:
        The output file path
    """
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    wb = Workbook()

    # Define styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # --- Sheet 1: Dashboard ---
    _build_dashboard_sheet(wb, results, header_font, header_fill, header_alignment, thin_border)

    # --- Sheet 2: Strong Buys ---
    _build_strong_buys_sheet(wb, results, header_font, header_fill, header_alignment, thin_border)

    # --- Sheet 3: Sector Summary ---
    _build_sector_summary_sheet(wb, results, header_font, header_fill, header_alignment, thin_border)

    # --- Sheet 4: Data Quality ---
    _build_data_quality_sheet(wb, results, header_font, header_fill, header_alignment, thin_border)

    # --- Sheet 5: Assumptions ---
    _build_assumptions_sheet(wb, header_font, header_fill, header_alignment, thin_border)

    # Save
    wb.save(output_path)
    logger.info(f"Excel report saved to {output_path}")
    return output_path


def _build_dashboard_sheet(wb, results, header_font, header_fill, header_alignment, thin_border):
    """Build the main dashboard sheet."""
    ws = wb.active
    ws.title = "Dashboard"

    # Define headers
    headers = [
        'Ticker', 'Company', 'Sector', 'Price', 'Intrinsic Value', 'Upside %',
        'Lower Band', 'Upper Band', 'Band Position', 'RSI', 'Signal', 'Conviction', 'Rationale',
        'P/E', 'P/B', 'EPS', 'ROA', 'ROE', 'EBITDA%', 'Net Margin%', 'D/E',
        'IV (Exit Mult)', 'IV (Perp Growth)', 'Exit Multiple', 'Implied g',
        'Return 1M', 'Return 6M', 'Return 1Y', 'Return 3Y', 'Std Dev 52W', 'Sharpe 52W',
        'Buy Price',
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Sort results by signal then conviction
    signal_order = {'STRONG BUY': 0, 'BUY': 1, 'HOLD': 2, 'SELL': 3, 'STRONG SELL': 4}

    def sort_key(r):
        sig = r.get('signal', 'HOLD')
        # Extract base signal (remove confidence suffix)
        base_sig = sig.split('(')[0].strip()
        return (signal_order.get(base_sig, 5), -(r.get('conviction', 0) or 0))

    sorted_results = sorted(results, key=sort_key)

    # Write data rows
    for row_idx, result in enumerate(sorted_results, start=2):
        ws.cell(row=row_idx, column=1).value = result.get('ticker', 'N/A')
        ws.cell(row=row_idx, column=2).value = result.get('ticker', 'N/A')  # Company not available
        ws.cell(row=row_idx, column=3).value = result.get('sector', 'N/A')

        # Price
        price_cell = ws.cell(row=row_idx, column=4)
        price = result.get('current_price')
        price_cell.value = price if price is not None else 'N/A'
        if isinstance(price, (int, float)):
            price_cell.number_format = '$#,##0.00'

        # Intrinsic Value
        iv_cell = ws.cell(row=row_idx, column=5)
        iv = result.get('intrinsic_value')
        iv_cell.value = iv if iv is not None else 'N/A'
        if isinstance(iv, (int, float)):
            iv_cell.number_format = '$#,##0.00'

        # Upside %
        upside_cell = ws.cell(row=row_idx, column=6)
        upside = result.get('upside_pct')
        upside_cell.value = upside if upside is not None else 'N/A'
        if isinstance(upside, (int, float)):
            upside_cell.number_format = '0.0%'

        # Lower Band
        lower_cell = ws.cell(row=row_idx, column=7)
        lower = result.get('tech_lower_band')
        lower_cell.value = lower if lower is not None else 'N/A'
        if isinstance(lower, (int, float)):
            lower_cell.number_format = '$#,##0.00'

        # Upper Band
        upper_cell = ws.cell(row=row_idx, column=8)
        upper = result.get('tech_upper_band')
        upper_cell.value = upper if upper is not None else 'N/A'
        if isinstance(upper, (int, float)):
            upper_cell.number_format = '$#,##0.00'

        # Band Position
        band_pos_cell = ws.cell(row=row_idx, column=9)
        band_pos = result.get('tech_band_position', 'N/A')
        band_pos_cell.value = band_pos

        # RSI
        rsi_cell = ws.cell(row=row_idx, column=10)
        rsi = result.get('tech_rsi')
        rsi_cell.value = rsi if rsi is not None else 'N/A'
        if isinstance(rsi, (int, float)):
            rsi_cell.number_format = '0'

        # Signal
        ws.cell(row=row_idx, column=11).value = result.get('signal', 'N/A')

        # Conviction
        conv_cell = ws.cell(row=row_idx, column=12)
        conv = result.get('conviction')
        conv_cell.value = conv if conv is not None else 'N/A'
        if isinstance(conv, (int, float)):
            conv_cell.number_format = '0'

        # Rationale
        ws.cell(row=row_idx, column=13).value = result.get('rationale', 'N/A')

        # P/E
        pe_cell = ws.cell(row=row_idx, column=14)
        pe = result.get('trailing_pe')
        pe_cell.value = pe if pe is not None else 'N/A'
        if isinstance(pe, (int, float)):
            pe_cell.number_format = '0.0'

        # P/B
        pb_cell = ws.cell(row=row_idx, column=15)
        pb = result.get('price_to_book')
        pb_cell.value = pb if pb is not None else 'N/A'
        if isinstance(pb, (int, float)):
            pb_cell.number_format = '0.00'

        # EPS
        eps_cell = ws.cell(row=row_idx, column=16)
        eps = result.get('trailing_eps')
        eps_cell.value = eps if eps is not None else 'N/A'
        if isinstance(eps, (int, float)):
            eps_cell.number_format = '$#,##0.00'

        # ROA
        roa_cell = ws.cell(row=row_idx, column=17)
        roa = result.get('return_on_assets')
        roa_cell.value = roa if roa is not None else 'N/A'
        if isinstance(roa, (int, float)):
            roa_cell.number_format = '0.0%'

        # ROE
        roe_cell = ws.cell(row=row_idx, column=18)
        roe = result.get('return_on_equity')
        roe_cell.value = roe if roe is not None else 'N/A'
        if isinstance(roe, (int, float)):
            roe_cell.number_format = '0.0%'

        # EBITDA%
        ebitda_cell = ws.cell(row=row_idx, column=19)
        ebitda = result.get('ebitda_margin')
        ebitda_cell.value = ebitda if ebitda is not None else 'N/A'
        if isinstance(ebitda, (int, float)):
            ebitda_cell.number_format = '0.0%'

        # Net Margin%
        net_margin_cell = ws.cell(row=row_idx, column=20)
        net_margin = result.get('net_margin')
        net_margin_cell.value = net_margin if net_margin is not None else 'N/A'
        if isinstance(net_margin, (int, float)):
            net_margin_cell.number_format = '0.0%'

        # D/E
        de_cell = ws.cell(row=row_idx, column=21)
        de = result.get('debt_to_equity')
        de_cell.value = de if de is not None else 'N/A'
        if isinstance(de, (int, float)):
            de_cell.number_format = '0.0'

        # IV (Exit Mult)
        iv_exit_cell = ws.cell(row=row_idx, column=22)
        iv_exit = result.get('iv_exit_multiple')
        iv_exit_cell.value = iv_exit if iv_exit is not None else 'N/A'
        if isinstance(iv_exit, (int, float)):
            iv_exit_cell.number_format = '$#,##0.00'

        # IV (Perp Growth)
        iv_perp_cell = ws.cell(row=row_idx, column=23)
        iv_perp = result.get('iv_perpetual_growth')
        iv_perp_cell.value = iv_perp if iv_perp is not None else 'N/A'
        if isinstance(iv_perp, (int, float)):
            iv_perp_cell.number_format = '$#,##0.00'

        # Exit Multiple
        exit_mult_cell = ws.cell(row=row_idx, column=24)
        exit_mult = result.get('exit_multiple_used')
        exit_mult_cell.value = exit_mult if exit_mult is not None else 'N/A'
        if isinstance(exit_mult, (int, float)):
            exit_mult_cell.number_format = '0.0x'

        # Implied g
        implied_g_cell = ws.cell(row=row_idx, column=25)
        implied_g = result.get('implied_perpetuity_growth')
        implied_g_cell.value = implied_g if implied_g is not None else 'N/A'
        if isinstance(implied_g, (int, float)):
            implied_g_cell.number_format = '0.0%'

        # Return 1M
        ret1m_cell = ws.cell(row=row_idx, column=26)
        ret1m = result.get('tech_return_1m')
        ret1m_cell.value = ret1m if ret1m is not None else 'N/A'
        if isinstance(ret1m, (int, float)):
            ret1m_cell.number_format = '0.0%'

        # Return 6M
        ret6m_cell = ws.cell(row=row_idx, column=27)
        ret6m = result.get('tech_return_6m')
        ret6m_cell.value = ret6m if ret6m is not None else 'N/A'
        if isinstance(ret6m, (int, float)):
            ret6m_cell.number_format = '0.0%'

        # Return 1Y
        ret1y_cell = ws.cell(row=row_idx, column=28)
        ret1y = result.get('tech_return_1y')
        ret1y_cell.value = ret1y if ret1y is not None else 'N/A'
        if isinstance(ret1y, (int, float)):
            ret1y_cell.number_format = '0.0%'

        # Return 3Y
        ret3y_cell = ws.cell(row=row_idx, column=29)
        ret3y = result.get('tech_return_3y')
        ret3y_cell.value = ret3y if ret3y is not None else 'N/A'
        if isinstance(ret3y, (int, float)):
            ret3y_cell.number_format = '0.0%'

        # Std Dev 52W
        stddev_cell = ws.cell(row=row_idx, column=30)
        stddev = result.get('tech_std_dev_52w')
        stddev_cell.value = stddev if stddev is not None else 'N/A'
        if isinstance(stddev, (int, float)):
            stddev_cell.number_format = '0.0%'

        # Sharpe 52W
        sharpe_cell = ws.cell(row=row_idx, column=31)
        sharpe = result.get('tech_sharpe_52w')
        sharpe_cell.value = sharpe if sharpe is not None else 'N/A'
        if isinstance(sharpe, (int, float)):
            sharpe_cell.number_format = '0.00'

        # Buy Price (IV * 0.75)
        buy_price_cell = ws.cell(row=row_idx, column=32)
        iv = result.get('intrinsic_value')
        if isinstance(iv, (int, float)) and iv > 0:
            buy_price = iv * 0.75
            buy_price_cell.value = buy_price
            buy_price_cell.number_format = '$#,##0.00'
        else:
            buy_price_cell.value = 'N/A'

    # Apply borders to all data cells
    for row in ws.iter_rows(min_row=2, max_row=len(sorted_results) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Conditional formatting on Upside % column (column F = 6)
    upside_col = get_column_letter(6)
    if len(sorted_results) > 0:
        range_str = f'{upside_col}2:{upside_col}{len(sorted_results) + 1}'

        # Dark green: > 0.30
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['0.30'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                      font=Font(color='006100'))
        )

        # Light green: 0.15 to 0.30
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='between', formula=['0.15', '0.30'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='D9F2E6', end_color='D9F2E6', fill_type='solid'),
                      font=Font(color='006100'))
        )

        # Light red: -0.25 to -0.10
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='between', formula=['-0.25', '-0.10'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                      font=Font(color='9C0006'))
        )

        # Dark red: < -0.25
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='lessThan', formula=['-0.25'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                      font=Font(color='9C0006', bold=True))
        )

    # Conditional formatting on Sharpe 52W column (column 31)
    sharpe_col = get_column_letter(31)
    if len(sorted_results) > 0:
        range_str = f'{sharpe_col}2:{sharpe_col}{len(sorted_results) + 1}'

        # Green: Sharpe > 1.0
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['1.0'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                      font=Font(color='006100'))
        )

        # Red: Sharpe < 0
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='lessThan', formula=['0'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                      font=Font(color='9C0006'))
        )

    # Conditional formatting on ROE column (column R = 18)
    roe_col = get_column_letter(18)
    if len(sorted_results) > 0:
        range_str = f'{roe_col}2:{roe_col}{len(sorted_results) + 1}'

        # Green fill: ROE > 15%
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['0.15'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                      font=Font(color='006100'))
        )

        # Red fill: ROE < 5%
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='lessThan', formula=['0.05'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                      font=Font(color='9C0006'))
        )

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    if len(sorted_results) > 0:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(sorted_results) + 1}'

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = max(len(str(header)), 12) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length


def _build_strong_buys_sheet(wb, results, header_font, header_fill, header_alignment, thin_border):
    """Build the Strong Buys sheet."""
    ws = wb.create_sheet(title="Strong Buys")

    # Filter for STRONG BUY and BUY signals
    filtered_results = [
        r for r in results
        if r.get('signal', '').startswith('STRONG BUY') or r.get('signal', '').startswith('BUY')
    ]

    # Sort by conviction desc, then upside_pct desc
    filtered_results.sort(
        key=lambda r: (-(r.get('conviction', 0) or 0), -(r.get('upside_pct', 0) or 0))
    )

    # Define headers (same as Dashboard)
    headers = [
        'Ticker', 'Company', 'Sector', 'Price', 'Intrinsic Value', 'Upside %',
        'Lower Band', 'Upper Band', 'Band Position', 'RSI', 'Signal', 'Conviction', 'Rationale',
        'P/E', 'P/B', 'EPS', 'ROA', 'ROE', 'EBITDA%', 'Net Margin%', 'D/E',
        'IV (Exit Mult)', 'IV (Perp Growth)', 'Exit Multiple', 'Implied g',
        'Return 1M', 'Return 6M', 'Return 1Y', 'Return 3Y', 'Std Dev 52W', 'Sharpe 52W',
        'Buy Price',
    ]

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, result in enumerate(filtered_results, start=2):
        ws.cell(row=row_idx, column=1).value = result.get('ticker', 'N/A')
        ws.cell(row=row_idx, column=2).value = result.get('ticker', 'N/A')
        ws.cell(row=row_idx, column=3).value = result.get('sector', 'N/A')

        # Price
        price_cell = ws.cell(row=row_idx, column=4)
        price = result.get('current_price')
        price_cell.value = price if price is not None else 'N/A'
        if isinstance(price, (int, float)):
            price_cell.number_format = '$#,##0.00'

        # Intrinsic Value
        iv_cell = ws.cell(row=row_idx, column=5)
        iv = result.get('intrinsic_value')
        iv_cell.value = iv if iv is not None else 'N/A'
        if isinstance(iv, (int, float)):
            iv_cell.number_format = '$#,##0.00'

        # Upside %
        upside_cell = ws.cell(row=row_idx, column=6)
        upside = result.get('upside_pct')
        upside_cell.value = upside if upside is not None else 'N/A'
        if isinstance(upside, (int, float)):
            upside_cell.number_format = '0.0%'

        # Lower Band
        lower_cell = ws.cell(row=row_idx, column=7)
        lower = result.get('tech_lower_band')
        lower_cell.value = lower if lower is not None else 'N/A'
        if isinstance(lower, (int, float)):
            lower_cell.number_format = '$#,##0.00'

        # Upper Band
        upper_cell = ws.cell(row=row_idx, column=8)
        upper = result.get('tech_upper_band')
        upper_cell.value = upper if upper is not None else 'N/A'
        if isinstance(upper, (int, float)):
            upper_cell.number_format = '$#,##0.00'

        # Band Position
        ws.cell(row=row_idx, column=9).value = result.get('tech_band_position', 'N/A')

        # RSI
        rsi_cell = ws.cell(row=row_idx, column=10)
        rsi = result.get('tech_rsi')
        rsi_cell.value = rsi if rsi is not None else 'N/A'
        if isinstance(rsi, (int, float)):
            rsi_cell.number_format = '0'

        # Signal
        ws.cell(row=row_idx, column=11).value = result.get('signal', 'N/A')

        # Conviction
        conv_cell = ws.cell(row=row_idx, column=12)
        conv = result.get('conviction')
        conv_cell.value = conv if conv is not None else 'N/A'
        if isinstance(conv, (int, float)):
            conv_cell.number_format = '0'

        # Rationale
        ws.cell(row=row_idx, column=13).value = result.get('rationale', 'N/A')

        # P/E
        pe_cell = ws.cell(row=row_idx, column=14)
        pe = result.get('trailing_pe')
        pe_cell.value = pe if pe is not None else 'N/A'
        if isinstance(pe, (int, float)):
            pe_cell.number_format = '0.0'

        # P/B
        pb_cell = ws.cell(row=row_idx, column=15)
        pb = result.get('price_to_book')
        pb_cell.value = pb if pb is not None else 'N/A'
        if isinstance(pb, (int, float)):
            pb_cell.number_format = '0.00'

        # EPS
        eps_cell = ws.cell(row=row_idx, column=16)
        eps = result.get('trailing_eps')
        eps_cell.value = eps if eps is not None else 'N/A'
        if isinstance(eps, (int, float)):
            eps_cell.number_format = '$#,##0.00'

        # ROA
        roa_cell = ws.cell(row=row_idx, column=17)
        roa = result.get('return_on_assets')
        roa_cell.value = roa if roa is not None else 'N/A'
        if isinstance(roa, (int, float)):
            roa_cell.number_format = '0.0%'

        # ROE
        roe_cell = ws.cell(row=row_idx, column=18)
        roe = result.get('return_on_equity')
        roe_cell.value = roe if roe is not None else 'N/A'
        if isinstance(roe, (int, float)):
            roe_cell.number_format = '0.0%'

        # EBITDA%
        ebitda_cell = ws.cell(row=row_idx, column=19)
        ebitda = result.get('ebitda_margin')
        ebitda_cell.value = ebitda if ebitda is not None else 'N/A'
        if isinstance(ebitda, (int, float)):
            ebitda_cell.number_format = '0.0%'

        # Net Margin%
        net_margin_cell = ws.cell(row=row_idx, column=20)
        net_margin = result.get('net_margin')
        net_margin_cell.value = net_margin if net_margin is not None else 'N/A'
        if isinstance(net_margin, (int, float)):
            net_margin_cell.number_format = '0.0%'

        # D/E
        de_cell = ws.cell(row=row_idx, column=21)
        de = result.get('debt_to_equity')
        de_cell.value = de if de is not None else 'N/A'
        if isinstance(de, (int, float)):
            de_cell.number_format = '0.0'

        # IV (Exit Mult)
        iv_exit_cell = ws.cell(row=row_idx, column=22)
        iv_exit = result.get('iv_exit_multiple')
        iv_exit_cell.value = iv_exit if iv_exit is not None else 'N/A'
        if isinstance(iv_exit, (int, float)):
            iv_exit_cell.number_format = '$#,##0.00'

        # IV (Perp Growth)
        iv_perp_cell = ws.cell(row=row_idx, column=23)
        iv_perp = result.get('iv_perpetual_growth')
        iv_perp_cell.value = iv_perp if iv_perp is not None else 'N/A'
        if isinstance(iv_perp, (int, float)):
            iv_perp_cell.number_format = '$#,##0.00'

        # Exit Multiple
        exit_mult_cell = ws.cell(row=row_idx, column=24)
        exit_mult = result.get('exit_multiple_used')
        exit_mult_cell.value = exit_mult if exit_mult is not None else 'N/A'
        if isinstance(exit_mult, (int, float)):
            exit_mult_cell.number_format = '0.0x'

        # Implied g
        implied_g_cell = ws.cell(row=row_idx, column=25)
        implied_g = result.get('implied_perpetuity_growth')
        implied_g_cell.value = implied_g if implied_g is not None else 'N/A'
        if isinstance(implied_g, (int, float)):
            implied_g_cell.number_format = '0.0%'

        # Return 1M
        ret1m_cell = ws.cell(row=row_idx, column=26)
        ret1m = result.get('tech_return_1m')
        ret1m_cell.value = ret1m if ret1m is not None else 'N/A'
        if isinstance(ret1m, (int, float)):
            ret1m_cell.number_format = '0.0%'

        # Return 6M
        ret6m_cell = ws.cell(row=row_idx, column=27)
        ret6m = result.get('tech_return_6m')
        ret6m_cell.value = ret6m if ret6m is not None else 'N/A'
        if isinstance(ret6m, (int, float)):
            ret6m_cell.number_format = '0.0%'

        # Return 1Y
        ret1y_cell = ws.cell(row=row_idx, column=28)
        ret1y = result.get('tech_return_1y')
        ret1y_cell.value = ret1y if ret1y is not None else 'N/A'
        if isinstance(ret1y, (int, float)):
            ret1y_cell.number_format = '0.0%'

        # Return 3Y
        ret3y_cell = ws.cell(row=row_idx, column=29)
        ret3y = result.get('tech_return_3y')
        ret3y_cell.value = ret3y if ret3y is not None else 'N/A'
        if isinstance(ret3y, (int, float)):
            ret3y_cell.number_format = '0.0%'

        # Std Dev 52W
        stddev_cell = ws.cell(row=row_idx, column=30)
        stddev = result.get('tech_std_dev_52w')
        stddev_cell.value = stddev if stddev is not None else 'N/A'
        if isinstance(stddev, (int, float)):
            stddev_cell.number_format = '0.0%'

        # Sharpe 52W
        sharpe_cell = ws.cell(row=row_idx, column=31)
        sharpe = result.get('tech_sharpe_52w')
        sharpe_cell.value = sharpe if sharpe is not None else 'N/A'
        if isinstance(sharpe, (int, float)):
            sharpe_cell.number_format = '0.00'

        # Buy Price (IV * 0.75)
        buy_price_cell = ws.cell(row=row_idx, column=32)
        iv = result.get('intrinsic_value')
        if isinstance(iv, (int, float)) and iv > 0:
            buy_price = iv * 0.75
            buy_price_cell.value = buy_price
            buy_price_cell.number_format = '$#,##0.00'
        else:
            buy_price_cell.value = 'N/A'

    # Apply borders
    for row in ws.iter_rows(min_row=2, max_row=len(filtered_results) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Conditional formatting on Upside %
    upside_col = get_column_letter(6)
    if len(filtered_results) > 0:
        range_str = f'{upside_col}2:{upside_col}{len(filtered_results) + 1}'

        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['0.30'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                      font=Font(color='006100'))
        )

        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='between', formula=['0.15', '0.30'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='D9F2E6', end_color='D9F2E6', fill_type='solid'),
                      font=Font(color='006100'))
        )

    # Conditional formatting on ROE column (column R = 18)
    roe_col = get_column_letter(18)
    if len(filtered_results) > 0:
        range_str = f'{roe_col}2:{roe_col}{len(filtered_results) + 1}'

        # Green fill: ROE > 15%
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['0.15'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                      font=Font(color='006100'))
        )

        # Red fill: ROE < 5%
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='lessThan', formula=['0.05'],
                      stopIfTrue=True,
                      fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                      font=Font(color='9C0006'))
        )

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    if len(filtered_results) > 0:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(filtered_results) + 1}'

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = max(len(str(header)), 12) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length


def _build_sector_summary_sheet(wb, results, header_font, header_fill, header_alignment, thin_border):
    """Build the Sector Summary sheet."""
    ws = wb.create_sheet(title="Sector Summary")

    # Group by sector
    sector_data = {}
    for result in results:
        sector = result.get('sector', 'Unknown')
        if sector not in sector_data:
            sector_data[sector] = {
                'count': 0,
                'upside_sum': 0,
                'upside_count': 0,
                'undervalued': 0,
                'overvalued': 0,
                'rsi_sum': 0,
                'rsi_count': 0,
            }

        sector_data[sector]['count'] += 1

        upside = result.get('upside_pct')
        if upside is not None:
            sector_data[sector]['upside_sum'] += upside
            sector_data[sector]['upside_count'] += 1

            if upside > 0.15:
                sector_data[sector]['undervalued'] += 1
            elif upside < -0.10:
                sector_data[sector]['overvalued'] += 1

        rsi = result.get('tech_rsi')
        if rsi is not None:
            sector_data[sector]['rsi_sum'] += rsi
            sector_data[sector]['rsi_count'] += 1

    # Define headers
    headers = ['Sector', '# Stocks', 'Avg Upside %', '# Undervalued', '# Overvalued', 'Avg RSI']

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    row_idx = 2
    for sector in sorted(sector_data.keys()):
        data = sector_data[sector]

        ws.cell(row=row_idx, column=1).value = sector
        ws.cell(row=row_idx, column=2).value = data['count']

        # Avg Upside %
        avg_upside_cell = ws.cell(row=row_idx, column=3)
        if data['upside_count'] > 0:
            avg_upside = data['upside_sum'] / data['upside_count']
            avg_upside_cell.value = avg_upside
            avg_upside_cell.number_format = '0.0%'
        else:
            avg_upside_cell.value = 'N/A'

        ws.cell(row=row_idx, column=4).value = data['undervalued']
        ws.cell(row=row_idx, column=5).value = data['overvalued']

        # Avg RSI
        avg_rsi_cell = ws.cell(row=row_idx, column=6)
        if data['rsi_count'] > 0:
            avg_rsi = data['rsi_sum'] / data['rsi_count']
            avg_rsi_cell.value = round(avg_rsi, 1)
            avg_rsi_cell.number_format = '0.0'
        else:
            avg_rsi_cell.value = 'N/A'

        row_idx += 1

    # Apply borders
    for row in ws.iter_rows(min_row=2, max_row=row_idx - 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    if len(sector_data) > 0:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_idx - 1}'

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = max(len(str(header)), 12) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length


def _build_data_quality_sheet(wb, results, header_font, header_fill, header_alignment, thin_border):
    """Build the Data Quality sheet."""
    ws = wb.create_sheet(title="Data Quality")

    # Define headers
    headers = ['Ticker', 'Confidence', 'Missing Fields', 'Last Updated', 'Notes']

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1).value = result.get('ticker', 'N/A')
        ws.cell(row=row_idx, column=2).value = result.get('confidence', 'N/A')

        # Identify missing fields
        important_fields = [
            'current_price', 'market_cap', 'sector', 'intrinsic_value',
            'upside_pct', 'tech_rsi', 'signal', 'conviction'
        ]
        missing = [field for field in important_fields if result.get(field) is None]
        missing_str = ', '.join(missing) if missing else 'None'
        ws.cell(row=row_idx, column=3).value = missing_str

        # Last Updated
        timestamp = result.get('fetch_timestamp')
        if timestamp:
            if isinstance(timestamp, str):
                ws.cell(row=row_idx, column=4).value = timestamp
            else:
                ws.cell(row=row_idx, column=4).value = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        else:
            ws.cell(row=row_idx, column=4).value = 'N/A'

        # Notes
        notes = []
        if result.get('dcf_note'):
            notes.append(f"DCF: {result['dcf_note']}")
        if result.get('valuation_warning'):
            notes.append(f"Warning: {result['valuation_warning']}")
        if result.get('relative_note'):
            notes.append(f"Relative: {result['relative_note']}")

        notes_str = ' | '.join(notes) if notes else 'None'
        ws.cell(row=row_idx, column=5).value = notes_str

    # Apply borders
    for row in ws.iter_rows(min_row=2, max_row=len(results) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    if len(results) > 0:
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(results) + 1}'

    # Auto-fit column widths
    for col_idx, header in enumerate(headers, start=1):
        if header == 'Notes':
            ws.column_dimensions[get_column_letter(col_idx)].width = 50
        elif header == 'Missing Fields':
            ws.column_dimensions[get_column_letter(col_idx)].width = 30
        else:
            max_length = max(len(str(header)), 12) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length


def _build_assumptions_sheet(wb, header_font, header_fill, header_alignment, thin_border):
    """Build the Assumptions sheet."""
    ws = wb.create_sheet(title="Assumptions")

    # Define headers
    headers = ['Parameter', 'Value', 'Description']

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Define assumptions data
    assumptions = [
        ('Report Generated', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC'), 'Timestamp of report generation'),
        ('Data Source', config.DATA_SOURCE, 'Source of financial data'),
        ('Cache Expiry', f'{config.CACHE_EXPIRY_HOURS} hours', 'Data cache validity period'),
        ('', '', ''),  # Blank row
        ('Risk-Free Rate', f'{config.RISK_FREE_RATE:.1%}', 'Used in WACC calculation'),
        ('Equity Risk Premium', f'{config.EQUITY_RISK_PREMIUM:.1%}', 'Market risk premium for WACC'),
        ('Default WACC', f'{config.DEFAULT_WACC:.1%}', 'Fallback discount rate'),
        ('DCF Projection Years', str(config.DCF_PROJECTION_YEARS), 'Number of years in DCF projection'),
        ('Margin of Safety', f'{config.MARGIN_OF_SAFETY:.0%}', 'Safety margin for buy price calculation'),
        ('', '', ''),  # Blank row
        ('--- Terminal Value ---', '', ''),
        ('Primary Method', 'Exit Multiple (EV/EBITDA)', 'Uses sector-specific EV/EBITDA multiples'),
        ('Secondary Method', 'Perpetual Growth (Gordon Growth)', 'Sanity check — sensitive to g and WACC'),
        ('Terminal Growth Rate', f'{config.TERMINAL_GROWTH_RATE:.1%}', 'Perpetual growth rate (secondary method)'),
        ('Default Exit Multiple', f'{config.DEFAULT_EXIT_MULTIPLE:.1f}x', 'Fallback when sector data unavailable'),
        ('Max Implied Growth Rate', f'{config.MAX_IMPLIED_GROWTH_RATE:.0%}', 'GDP ceiling — warns if implied g exceeds this'),
        ('', '', ''),  # Blank row
        ('--- Sector Exit Multiples ---', '', ''),
    ]

    # Add sector exit multiples
    for sector_name in sorted(config.SECTOR_EXIT_MULTIPLES.keys()):
        mult = config.SECTOR_EXIT_MULTIPLES[sector_name]
        assumptions.append((f'  {sector_name}', f'{mult:.1f}x', 'Default EV/EBITDA if peer median unavailable'))

    assumptions += [
        ('', '', ''),  # Blank row
        ('--- Technical Analysis ---', '', ''),
        ('Bollinger Window', str(config.BOLLINGER_WINDOW), 'Days for Bollinger Bands calculation'),
        ('Bollinger Std Dev', str(config.BOLLINGER_STD_DEV), 'Standard deviations for bands'),
        ('RSI Period', str(config.RSI_PERIOD), 'Period for RSI calculation'),
        ('', '', ''),  # Blank row
        ('--- Signal Thresholds ---', '', ''),
        ('Strong Buy Threshold', f'{config.STRONG_BUY_UPSIDE:.0%}', 'Minimum upside for STRONG BUY signal'),
        ('Buy Threshold', f'{config.BUY_UPSIDE:.0%}', 'Minimum upside for BUY signal'),
        ('Sell Threshold', f'{config.SELL_DOWNSIDE:.0%}', 'Maximum downside for SELL signal'),
        ('Strong Sell Threshold', f'{config.STRONG_SELL_DOWNSIDE:.0%}', 'Maximum downside for STRONG SELL signal'),
    ]

    # Write data rows
    for row_idx, (param, value, desc) in enumerate(assumptions, start=2):
        ws.cell(row=row_idx, column=1).value = param
        ws.cell(row=row_idx, column=2).value = value
        ws.cell(row=row_idx, column=3).value = desc

    # Apply borders to non-blank rows
    for row in ws.iter_rows(min_row=2, max_row=len(assumptions) + 1, min_col=1, max_col=len(headers)):
        if row[0].value:  # Only apply border if first column is not empty
            for cell in row:
                cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-fit column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
